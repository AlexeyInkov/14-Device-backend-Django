import csv
import os

from chardet.universaldetector import UniversalDetector
from django.core.files.uploadedfile import InMemoryUploadedFile
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill, Border, Side
from openpyxl.workbook import Workbook

from config import settings


def handle_uploaded_file(f: InMemoryUploadedFile) -> None:
    """Загрузка файла"""
    filename = os.path.join(settings.FILE_UPLOAD_DIR, f.name)
    with open(filename, "wb+") as destination:
        for chunk in f.chunks():
            destination.write(chunk)


def get_file_encoding(file_path: str) -> str:
    """Получение кодировки файла"""
    with open(file_path, "rb") as f:
        detector = UniversalDetector()
        detector.reset()
        for line in f.readlines():
            detector.feed(line)
        detector.close()
        result = detector.result["encoding"]
        if result == "windows-1251":
            return "cp1251"
        return result


def check_csv_file(file_path: str, fieldnames: list, encoding: str) -> bool:
    """Проверка заголовков таблицы csv файла"""
    with open(file_path, "r", encoding=encoding) as file:
        reader = csv.DictReader(file, delimiter=";")
        for field in fieldnames:
            if field not in reader.fieldnames:
                return False
        return True


def create_excel_from_dict_list(
    dict_list: list, output_filename: str, sheet_name="Sheet1"
) -> os.path:

    filepath = os.path.join(settings.FILE_UPLOAD_DIR, output_filename)

    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Записываем данные из списка словарей в Excel
    if dict_list:
        header = list(dict_list[0].keys())
        ws.append(header)  # Записываем заголовки

        for row in dict_list:
            ws.append([row[col] for col in header])

    # Настраиваем стили для красивого вида
    header_style = NamedStyle(name="header")
    header_style.font = Font(bold=True, color="FFFFFF")
    header_style.alignment = Alignment(horizontal="center", vertical="center")
    header_style.fill = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )
    border_style = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    header_style.border = border_style

    cell_style = NamedStyle(name="cell")
    cell_style.alignment = Alignment(horizontal="left", vertical="center")
    cell_style.border = border_style

    date_style = NamedStyle(name="data")
    date_style.number_format = "DD.MM.YYYY"
    date_style.alignment = Alignment(horizontal="left", vertical="center")
    date_style.border = border_style

    for cell in ws[1]:  # Применяем стиль к заголовкам
        cell.style = header_style

    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            if cell.is_date:
                cell.style = date_style
            else:
                cell.style = cell_style

    # Автоматическое изменение ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    # Сохраняем файл
    wb.save(filepath)
    return filepath
